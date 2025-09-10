"""
엑셀/CSV 파일 처리 서비스
"""
import pandas as pd
import openpyxl
from typing import List, Dict, Any, Tuple
import logging
from io import BytesIO
import json

logger = logging.getLogger(__name__)


class ExcelProcessingService:
    """엑셀/CSV 파일 처리 서비스"""
    
    def __init__(self):
        self.max_rows_per_chunk = 50  # 청크당 최대 행 수
        self.max_cell_length = 1000   # 셀당 최대 문자 수
    
    def process_excel_file(self, file_content: bytes, filename: str) -> List[Dict[str, Any]]:
        """
        XLSX 파일을 처리하여 청크 리스트 반환
        
        Args:
            file_content: 파일 바이트 데이터
            filename: 파일명
            
        Returns:
            청크 리스트 (각 청크는 content, chunk_type, metadata 포함)
        """
        try:
            chunks = []
            
            # openpyxl로 워크북 읽기
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 시트별 청크 생성
                sheet_chunks = self._process_sheet(sheet, sheet_name, filename)
                chunks.extend(sheet_chunks)
            
            logger.info(f"XLSX 파일 처리 완료: {filename}, 총 {len(chunks)}개 청크 생성")
            return chunks
            
        except Exception as e:
            logger.error(f"XLSX 파일 처리 실패: {filename}, 오류: {str(e)}")
            raise
    
    def process_csv_file(self, file_content: bytes, filename: str) -> List[Dict[str, Any]]:
        """
        CSV 파일을 처리하여 청크 리스트 반환
        
        Args:
            file_content: 파일 바이트 데이터
            filename: 파일명
            
        Returns:
            청크 리스트 (각 청크는 content, chunk_type, metadata 포함)
        """
        try:
            chunks = []
            
            # pandas로 CSV 읽기 (인코딩 자동 감지)
            encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(BytesIO(file_content), encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise ValueError("CSV 파일 인코딩을 감지할 수 없습니다.")
            
            # CSV 청크 생성
            csv_chunks = self._process_dataframe(df, filename, "CSV")
            chunks.extend(csv_chunks)
            
            logger.info(f"CSV 파일 처리 완료: {filename}, 총 {len(chunks)}개 청크 생성")
            return chunks
            
        except Exception as e:
            logger.error(f"CSV 파일 처리 실패: {filename}, 오류: {str(e)}")
            raise
    
    def _process_sheet(self, sheet, sheet_name: str, filename: str) -> List[Dict[str, Any]]:
        """시트를 처리하여 청크 리스트 생성"""
        chunks = []
        
        # 시트 정보 추출
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row <= 1:  # 헤더만 있거나 빈 시트
            return chunks
        
        # 헤더 행 추출
        headers = []
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column_{col}")
        
        # 데이터를 청크 단위로 처리
        current_chunk_rows = []
        current_chunk_index = 0
        
        for row_num in range(2, max_row + 1):  # 헤더 제외하고 시작
            row_data = []
            has_data = False
            
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row_num, column=col).value
                if cell_value is not None:
                    cell_str = str(cell_value).strip()
                    if len(cell_str) > self.max_cell_length:
                        cell_str = cell_str[:self.max_cell_length] + "..."
                    row_data.append(cell_str)
                    has_data = True
                else:
                    row_data.append("")
            
            if has_data:
                current_chunk_rows.append((row_num, row_data))
            
            # 청크 크기에 도달하거나 마지막 행인 경우
            if len(current_chunk_rows) >= self.max_rows_per_chunk or row_num == max_row:
                if current_chunk_rows:
                    chunk_content = self._create_chunk_content(headers, current_chunk_rows, sheet_name)
                    
                    chunk = {
                        'content': chunk_content,
                        'chunk_type': 'excel_sheet',
                        'chunk_metadata': {
                            'sheet_name': sheet_name,
                            'filename': filename,
                            'chunk_index': current_chunk_index,
                            'row_range': f"{current_chunk_rows[0][0]}-{current_chunk_rows[-1][0]}",
                            'total_columns': len(headers),
                            'rows_in_chunk': len(current_chunk_rows)
                        }
                    }
                    chunks.append(chunk)
                    current_chunk_index += 1
                    current_chunk_rows = []
        
        return chunks
    
    def _process_dataframe(self, df: pd.DataFrame, filename: str, file_type: str) -> List[Dict[str, Any]]:
        """DataFrame을 처리하여 청크 리스트 생성"""
        chunks = []
        
        if df.empty:
            return chunks
        
        # 컬럼명 정리
        headers = [str(col).strip() for col in df.columns]
        
        # 데이터를 청크 단위로 처리
        current_chunk_rows = []
        current_chunk_index = 0
        
        for idx, row in df.iterrows():
            row_data = []
            has_data = False
            
            for value in row:
                if pd.notna(value):
                    cell_str = str(value).strip()
                    if len(cell_str) > self.max_cell_length:
                        cell_str = cell_str[:self.max_cell_length] + "..."
                    row_data.append(cell_str)
                    has_data = True
                else:
                    row_data.append("")
            
            if has_data:
                current_chunk_rows.append((idx + 2, row_data))  # +2는 헤더 행 고려
            
            # 청크 크기에 도달하거나 마지막 행인 경우
            if len(current_chunk_rows) >= self.max_rows_per_chunk or idx == len(df) - 1:
                if current_chunk_rows:
                    chunk_content = self._create_chunk_content(headers, current_chunk_rows, file_type)
                    
                    chunk = {
                        'content': chunk_content,
                        'chunk_type': 'excel_sheet',
                        'chunk_metadata': {
                            'sheet_name': file_type,
                            'filename': filename,
                            'chunk_index': current_chunk_index,
                            'row_range': f"{current_chunk_rows[0][0]}-{current_chunk_rows[-1][0]}",
                            'total_columns': len(headers),
                            'rows_in_chunk': len(current_chunk_rows)
                        }
                    }
                    chunks.append(chunk)
                    current_chunk_index += 1
                    current_chunk_rows = []
        
        return chunks
    
    def _create_chunk_content(self, headers: List[str], rows: List[Tuple[int, List[str]]], sheet_name: str) -> str:
        """청크 내용을 구조화된 텍스트로 생성"""
        content_parts = []
        
        # 시트/파일 정보
        content_parts.append(f"=== {sheet_name} 데이터 ===")
        
        # 헤더 정보
        content_parts.append(f"컬럼: {', '.join(headers)}")
        content_parts.append("")
        
        # 데이터 행들
        for row_num, row_data in rows:
            row_text = f"행 {row_num}: "
            row_values = []
            
            for i, (header, value) in enumerate(zip(headers, row_data)):
                if value:  # 빈 값이 아닌 경우만
                    row_values.append(f"{header}={value}")
            
            if row_values:
                row_text += " | ".join(row_values)
                content_parts.append(row_text)
        
        return "\n".join(content_parts)
    
    def get_file_info(self, file_content: bytes, filename: str, mime_type: str) -> Dict[str, Any]:
        """파일 정보 추출"""
        try:
            if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                # XLSX 파일 정보
                workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
                sheet_names = workbook.sheetnames
                total_sheets = len(sheet_names)
                
                # 첫 번째 시트의 크기 정보
                first_sheet = workbook[sheet_names[0]]
                max_row = first_sheet.max_row
                max_col = first_sheet.max_column
                
                return {
                    'file_type': 'xlsx',
                    'total_sheets': total_sheets,
                    'sheet_names': sheet_names,
                    'max_rows': max_row,
                    'max_columns': max_col
                }
                
            elif mime_type == 'text/csv':
                # CSV 파일 정보
                encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
                df = None
                used_encoding = None
                
                for encoding in encodings:
                    try:
                        df = pd.read_csv(BytesIO(file_content), encoding=encoding)
                        used_encoding = encoding
                        break
                    except UnicodeDecodeError:
                        continue
                
                if df is None:
                    raise ValueError("CSV 파일 인코딩을 감지할 수 없습니다.")
                
                return {
                    'file_type': 'csv',
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'column_names': df.columns.tolist(),
                    'encoding': used_encoding
                }
            
            return {}
            
        except Exception as e:
            logger.error(f"파일 정보 추출 실패: {filename}, 오류: {str(e)}")
            return {}
